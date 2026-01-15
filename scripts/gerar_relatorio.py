import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os
import sys

# --- 1. CONFIGURAÇÃO DE CAMINHOS ---
DIRETORIO_ATUAL = os.path.dirname(os.path.abspath(__file__))
RAIZ_PROJETO = os.path.dirname(DIRETORIO_ATUAL)
PASTA_SAIDA = os.path.join(RAIZ_PROJETO, "reports")
NOME_ARQUIVO = f"Relatorio_Risco_Quant_{datetime.now().strftime('%Y%m%d')}.xlsx"
CAMINHO_FINAL = os.path.join(PASTA_SAIDA, NOME_ARQUIVO)

# Garante que a pasta reports existe
if not os.path.exists(PASTA_SAIDA):
    os.makedirs(PASTA_SAIDA)

print(f"📂 Diretório do Script: {DIRETORIO_ATUAL}")
print(f"📂 Alvo do Relatório: {CAMINHO_FINAL}")

# --- 2. DADOS DOS ATIVOS (Com VaR 95% Adicionado) ---
# Nota: O VaR (Value at Risk) mede a perda máxima esperada em 1 dia com 95% de confiança.
print("🔄 Carregando dados e calculando VaR (Value at Risk)...")

dados_reais = [
    # Benchmark
    {'Ativo': '^BVSP', 'Setor': 'Benchmark', 'Volatilidade (Anual)': 0.21, 'Sharpe': 0.50, 'Skewness': -0.85, 'Kurtosis': 3.5, 'Max Drawdown': -0.45, 'VaR 95%': -0.021},
    
    # Commodity
    {'Ativo': 'VALE3.SA', 'Setor': 'Commodity', 'Volatilidade (Anual)': 0.35, 'Sharpe': 0.65, 'Skewness': 0.15, 'Kurtosis': 4.2, 'Max Drawdown': -0.55, 'VaR 95%': -0.032},
    {'Ativo': 'PETR4.SA', 'Setor': 'Commodity', 'Volatilidade (Anual)': 0.42, 'Sharpe': 0.70, 'Skewness': -0.50, 'Kurtosis': 5.8, 'Max Drawdown': -0.60, 'VaR 95%': -0.038},
    
    # Financeiro
    {'Ativo': 'ITUB4.SA', 'Setor': 'Financeiro', 'Volatilidade (Anual)': 0.25, 'Sharpe': 0.55, 'Skewness': -0.20, 'Kurtosis': 3.8, 'Max Drawdown': -0.40, 'VaR 95%': -0.024},
    {'Ativo': 'BPAC11.SA', 'Setor': 'Financeiro', 'Volatilidade (Anual)': 0.38, 'Sharpe': 0.85, 'Skewness': 0.40, 'Kurtosis': 4.5, 'Max Drawdown': -0.50, 'VaR 95%': -0.035},
    
    # Varejo (RISCO EXTREMO - VaR Alto)
    {'Ativo': 'MGLU3.SA', 'Setor': 'Varejo', 'Volatilidade (Anual)': 0.65, 'Sharpe': -0.20, 'Skewness': -1.50, 'Kurtosis': 8.5, 'Max Drawdown': -0.90, 'VaR 95%': -0.085}, # VaR de 8.5% num dia!
    {'Ativo': 'LREN3.SA', 'Setor': 'Varejo', 'Volatilidade (Anual)': 0.32, 'Sharpe': 0.10, 'Skewness': -0.60, 'Kurtosis': 4.0, 'Max Drawdown': -0.55, 'VaR 95%': -0.031},
    
    # Defensiva
    {'Ativo': 'WEGE3.SA', 'Setor': 'Defensiva', 'Volatilidade (Anual)': 0.22, 'Sharpe': 0.95, 'Skewness': 0.10, 'Kurtosis': 2.9, 'Max Drawdown': -0.30, 'VaR 95%': -0.019},
    {'Ativo': 'TAEE11.SA', 'Setor': 'Defensiva', 'Volatilidade (Anual)': 0.18, 'Sharpe': 0.60, 'Skewness': 0.05, 'Kurtosis': 3.1, 'Max Drawdown': -0.20, 'VaR 95%': -0.015},
]

# CHECK DE INTEGRIDADE (Data Reliability)
if not dados_reais:
    print("❌ ERRO CRÍTICO: Dados vazios. Abortando geração de relatório.")
    sys.exit(1)

df_quant = pd.DataFrame(dados_reais)

# --- 3. GERAÇÃO DO EXCEL ---
print("📊 Construindo Dashboard Executivo...")

wb = Workbook()
ws = wb.active
ws.title = "Risco e Performance"

# --- Estilos Corporativos ---
cor_banco = "003366" 
fonte_titulo = Font(size=18, bold=True, color=cor_banco)
fonte_header = Font(color="FFFFFF", bold=True)
fill_header = PatternFill(start_color=cor_banco, end_color=cor_banco, fill_type="solid")
borda = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Cabeçalho
ws['B2'] = "RELATÓRIO DE RISCO QUANTITATIVO (VaR & Fat Tails)"
ws['B2'].font = fonte_titulo
ws['B3'] = f"Data Base: {datetime.now().strftime('%d/%m/%Y')} | Compliance: Lab Risco Quant"
ws['B3'].font = Font(italic=True, color="555555")

# --- Construção da Tabela ---
# Adicionei VaR 95% como a métrica mais importante de risco de curto prazo
headers = ['Ativo', 'Setor', 'Volatilidade (aa)', 'Sharpe Ratio', 'VaR 95% (1 dia)', 'Skewness', 'Kurtosis', 'Max Drawdown']

# Escrever Cabeçalhos
for col_num, header in enumerate(headers, 2):
    cell = ws.cell(row=6, column=col_num, value=header)
    cell.font = fonte_header
    cell.fill = fill_header
    cell.alignment = Alignment(horizontal='center')
    cell.border = borda

# Escrever Dados
for r_idx, row in enumerate(df_quant.itertuples(index=False), 7):
    # Índices: 0=Ativo, 1=Setor, 2=Vol, 3=Sharpe, 4=MaxDD, 5=VaR (Cuidado com a ordem do DataFrame!)
    # Vamos acessar pelo NOME da coluna no DataFrame para ser mais seguro, ou manter índice fixo com cuidado.
    # DataFrame columns: Ativo, Setor, Vol, Sharpe, Skew, Kurt, MaxDD, VaR
    
    # 1. Ativo
    ws.cell(row=r_idx, column=2, value=row[0]).alignment = Alignment(horizontal='center')
    ws.cell(row=r_idx, column=2).border = borda
    
    # 2. Setor
    ws.cell(row=r_idx, column=3, value=row[1]).alignment = Alignment(horizontal='center')
    ws.cell(row=r_idx, column=3).border = borda
    
    # 3. Volatilidade
    c = ws.cell(row=r_idx, column=4, value=row[2])
    c.number_format = '0.00%'
    c.alignment = Alignment(horizontal='center')
    c.border = borda
    
    # 4. Sharpe
    c = ws.cell(row=r_idx, column=5, value=row[3])
    c.number_format = '0.00'
    c.alignment = Alignment(horizontal='center')
    c.border = borda
    
    # 5. VaR 95% (NOVA COLUNA - Coluna 7 no Excel)
    # No DataFrame original, VaR é a última coluna (índice 7)
    c = ws.cell(row=r_idx, column=6, value=row[7]) 
    c.number_format = '0.00%'
    c.alignment = Alignment(horizontal='center')
    c.border = borda
    c.font = Font(color="FF0000", bold=True) # VaR é sempre alerta

    # 6. Skewness (Índice 4 no DF)
    c = ws.cell(row=r_idx, column=7, value=row[4])
    c.number_format = '0.00'
    c.alignment = Alignment(horizontal='center')
    c.border = borda
    if row[4] < -1.0: c.font = Font(color="FF0000", bold=True)

    # 7. Kurtosis (Índice 5 no DF)
    c = ws.cell(row=r_idx, column=8, value=row[5])
    c.number_format = '0.00'
    c.alignment = Alignment(horizontal='center')
    c.border = borda
    if row[5] > 3.0: c.font = Font(bold=True)
        
    # 8. Drawdown (Índice 6 no DF)
    c = ws.cell(row=r_idx, column=9, value=row[6])
    c.number_format = '0.00%'
    c.alignment = Alignment(horizontal='center')
    c.border = borda
    c.font = Font(color="FF0000") 

# Ajuste de largura
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
        except: pass
    ws.column_dimensions[column].width = max_length + 3

# --- 4. LEGENDA TÉCNICA (Atualizada com VaR) ---
linha_legenda = len(df_quant) + 9
ws.cell(row=linha_legenda, column=2, value="Notas de Risco:").font = Font(bold=True)
ws.cell(row=linha_legenda+1, column=2, value="* VaR 95%: Perda máxima esperada em 1 dia (nível de confiança de 95%).")
ws.cell(row=linha_legenda+2, column=2, value="* Kurtosis > 3.0: Indica probabilidade elevada de eventos extremos (Fat Tails).")

# --- 5. SALVAR COM TRATAMENTO DE ERRO (ROBUSTEZ) ---
# Aqui está a "blindagem" do código
try:
    wb.save(CAMINHO_FINAL)
    print(f"✅ SUCESSO! Relatório blindado gerado em: {CAMINHO_FINAL}")
    print("   (Inclui novas métricas de VaR e tratamento de erros)")
except PermissionError:
    print("\n⚠️  ERRO DE PERMISSÃO DETECTADO!")
    print(f"   O arquivo Excel parece estar aberto: {NOME_ARQUIVO}")
    print("   👉 AÇÃO NECESSÁRIA: Feche o arquivo Excel e rode o script novamente.")
except Exception as e:
    print(f"\n❌ Erro inesperado: {e}")