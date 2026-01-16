import pandas as pd
import numpy as np
from scipy.stats import skew, kurtosis
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.formatting.rule import DataBarRule
import sqlite3
import os
import sys
from datetime import datetime

# --- 1. CONFIGURAÇÃO ---
DIRETORIO_ATUAL = os.path.dirname(os.path.abspath(__file__))
RAIZ_PROJETO = os.path.dirname(DIRETORIO_ATUAL)
CAMINHO_DB = os.path.join(RAIZ_PROJETO, "dados", "mercado.db")
PASTA_SAIDA = os.path.join(RAIZ_PROJETO, "reports")
NOME_ARQUIVO = f"Relatorio_Risco_Quant_{datetime.now().strftime('%Y%m%d')}.xlsx"
CAMINHO_FINAL = os.path.join(PASTA_SAIDA, NOME_ARQUIVO)

MAPA_SETORES = {
    '^BVSP': 'Benchmark',
    'VALE3.SA': 'Commodity', 'PETR4.SA': 'Commodity',
    'ITUB4.SA': 'Financeiro', 'BPAC11.SA': 'Financeiro',
    'MGLU3.SA': 'Varejo', 'LREN3.SA': 'Varejo',
    'WEGE3.SA': 'Defensiva', 'TAEE11.SA': 'Defensiva'
}
RISK_FREE = 0.1075

if not os.path.exists(PASTA_SAIDA): os.makedirs(PASTA_SAIDA)

# --- 2. MOTOR DE CÁLCULO (SQL) ---
def calcular_metricas_sql():
    print(f"🔌 Lendo SQL: {CAMINHO_DB}")
    if not os.path.exists(CAMINHO_DB):
        print("❌ Erro: DB não encontrado. Rode etl_sql.py.")
        sys.exit(1)
    conn = sqlite3.connect(CAMINHO_DB)
    try:
        df_precos = pd.read_sql("SELECT * FROM cotacoes", conn, index_col='Date')
        df_precos.index = pd.to_datetime(df_precos.index)
        retornos = df_precos.pct_change().dropna()
        
        lista = []
        for ticker in retornos.columns:
            if ticker not in MAPA_SETORES: continue
            r = retornos[ticker]
            vol = r.std() * np.sqrt(252)
            sharpe = (r.mean()*252 - RISK_FREE) / vol
            var95 = np.percentile(r, 5)
            lista.append({
                'Ativo': ticker, 'Setor': MAPA_SETORES.get(ticker),
                'Volatilidade': vol, 'Sharpe': sharpe, 'VaR 95%': var95,
                'Skewness': skew(r), 'Kurtosis': kurtosis(r, fisher=False),
                'MaxDrawdown': ((1+r).cumprod()/ (1+r).cumprod().cummax() - 1).min()
            })
        return pd.DataFrame(lista)
    except Exception as e: print(f"❌ Erro SQL: {e}"); return pd.DataFrame()
    finally: conn.close()

# --- 3. GERAÇÃO DO EXCEL ---
df = calcular_metricas_sql()
if df.empty: sys.exit(1)

print(f"📊 Gerando Excel Visual em: {CAMINHO_FINAL}")
wb = Workbook()
ws = wb.active
ws.title = "Dashboard Risco"

# Estilos
cor_azul = "003366"
cor_vermelha = "C00000"
font_titulo = Font(size=16, bold=True, color=cor_azul)
font_header = Font(color="FFFFFF", bold=True)
fill_header = PatternFill(start_color=cor_azul, end_color=cor_azul, fill_type="solid")
align_center = Alignment(horizontal='center', vertical='center')
borda_fina = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Títulos
ws['B2'] = "MONITOR DE RISCO QUANTITATIVO (SQL BACKEND)"
ws['B2'].font = font_titulo
ws['B3'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

# Tabela Principal
cols = ['Ativo', 'Setor', 'Volatilidade (aa)', 'Sharpe Ratio', 'VaR 95% (1d)', 'Skewness', 'Kurtosis', 'Max Drawdown']
ws.append([]); ws.append([]); ws.append([''] + cols) # Linha 5

for i in range(2, 10):
    cell = ws.cell(row=5, column=i)
    cell.fill = fill_header; cell.font = font_header; cell.alignment = align_center; cell.border = borda_fina

num_linhas = len(df)
for r_idx, row in enumerate(df.itertuples(index=False), 6):
    vals = [row.Ativo, row.Setor, row.Volatilidade, row.Sharpe, row._4, row.Skewness, row.Kurtosis, row.MaxDrawdown]
    ws.append([''] + vals)
    
    for c_idx in range(2, 10):
        cell = ws.cell(row=r_idx, column=c_idx)
        cell.alignment = align_center; cell.border = borda_fina
        if c_idx in [4, 6, 9]: cell.number_format = '0.00%' 
        else: cell.number_format = '0.00'

    if row._4 < -0.05: ws.cell(row=r_idx, column=6).font = Font(color=cor_vermelha, bold=True) 
    if row.Kurtosis > 3.0: ws.cell(row=r_idx, column=8).font = Font(bold=True) 

# --- CORREÇÃO AQUI (AS LINHAS QUE DAVAM ERRO) ---
# Adiciona barras de dados coloridas (Vermelho para VaR, Azul para Sharpe)
rule_var = DataBarRule(start_type='min', end_type='max', color="FF6384")
rule_sharpe = DataBarRule(start_type='min', end_type='max', color="36A2EB")

ws.conditional_formatting.add(f'F6:F{5+num_linhas}', rule_var)
ws.conditional_formatting.add(f'E6:E{5+num_linhas}', rule_sharpe)

for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = 16

# --- GRÁFICO DE DISPERSÃO ---
chart = ScatterChart()
chart.title = "Risco (Volatilidade) x Retorno Ajustado (Sharpe)"
chart.style = 2
chart.x_axis.title = 'Volatilidade Anualizada'
chart.y_axis.title = 'Índice de Sharpe'

xvalues = Reference(ws, min_col=4, min_row=6, max_row=5+num_linhas)
yvalues = Reference(ws, min_col=5, min_row=6, max_row=5+num_linhas)
series = Series(yvalues, xvalues, title_from_data=False)
series.marker.symbol = "circle"
series.marker.graphicalProperties.solidFill = "003366"
chart.series.append(series)
ws.add_chart(chart, "K2") 

# --- 4. GLOSSÁRIO TÉCNICO ---
linha_glossario = 5 + num_linhas + 4

ws.cell(row=linha_glossario, column=2, value="GLOSSÁRIO DE MÉTRICAS").font = Font(bold=True, size=12, color=cor_azul)

definicoes = [
    ("Sharpe Ratio", "Retorno excedente por unidade de risco. Quanto maior, melhor a eficiência do ativo."),
    ("VaR 95% (Value at Risk)", "Perda máxima esperada em 1 dia (95% confiança). Ex: -2% = chance de cair mais que isso é 5%."),
    ("Volatilidade", "Mede a intensidade das oscilações do preço (Risco total)."),
    ("Skewness (Assimetria)", "Mede a inclinação da curva. Negativo = maior frequência de quedas abruptas."),
    ("Kurtosis (Curtose)", "Detector de Caudas Gordas. > 3 indica alta probabilidade de eventos extremos."),
    ("Max Drawdown", "A pior queda percentual registrada do topo histórico até o fundo.")
]

for i, (termo, definicao) in enumerate(definicoes):
    linha_atual = linha_glossario + 1 + i
    c_termo = ws.cell(row=linha_atual, column=2, value=termo)
    c_termo.font = Font(bold=True)
    c_termo.border = borda_fina
    
    c_def = ws.cell(row=linha_atual, column=3, value=definicao)
    c_def.border = borda_fina
    ws.merge_cells(start_row=linha_atual, start_column=3, end_row=linha_atual, end_column=9)

try:
    wb.save(CAMINHO_FINAL)
    print("✅ RELATÓRIO COMPLETO GERADO COM SUCESSO!")
except PermissionError: print("⚠️ Erro: Feche o Excel.")