import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import ScatterChart, BarChart, Series, Reference
from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter
import sqlite3
import os
import sys
from datetime import datetime

# ==============================================================================
# 1. IMPORTAÇÃO SEGURA
# ==============================================================================
try:
    import dados_mercado as dm
    print("✅ Biblioteca 'dados_mercado' carregada com sucesso!")
except ImportError:
    print("❌ ERRO CRÍTICO: 'dados_mercado.py' não encontrado na pasta.")
    sys.exit(1)

# ==============================================================================
# 2. CONFIGURAÇÃO
# ==============================================================================
DIRETORIO_ATUAL = os.path.dirname(os.path.abspath(__file__))
RAIZ_PROJETO = os.path.dirname(os.path.dirname(DIRETORIO_ATUAL))

CAMINHO_DB = os.path.join(RAIZ_PROJETO, "dados", "mercado.db")
if not os.path.exists(CAMINHO_DB): CAMINHO_DB = "mercado.db"

PASTA_SAIDA = os.path.join(RAIZ_PROJETO, "reports")
NOME_ARQUIVO = f"Relatorio_Risco_Final_{datetime.now().strftime('%Y%m%d')}.xlsx"
CAMINHO_FINAL = os.path.join(PASTA_SAIDA, NOME_ARQUIVO)

MAPA_SETORES = {'^BVSP': 'Benchmark', 'VALE3.SA': 'Commodity', 'PETR4.SA': 'Commodity', 'ITUB4.SA': 'Financeiro', 'BPAC11.SA': 'Financeiro', 'MGLU3.SA': 'Varejo', 'LREN3.SA': 'Varejo', 'WEGE3.SA': 'Defensiva', 'TAEE11.SA': 'Defensiva'}
RISK_FREE = 0.1075

if not os.path.exists(PASTA_SAIDA): os.makedirs(PASTA_SAIDA)

# ==============================================================================
# 3. MOTOR DE CÁLCULO
# ==============================================================================
def calcular_metricas_sql():
    if not os.path.exists(CAMINHO_DB): return pd.DataFrame()
    conn = sqlite3.connect(CAMINHO_DB)
    try:
        df_precos = pd.read_sql("SELECT * FROM cotacoes", conn, index_col='Date')
        df_precos.index = pd.to_datetime(df_precos.index)
        retornos = df_precos.pct_change().dropna()
        bench_ret = retornos['^BVSP'] if '^BVSP' in retornos.columns else pd.Series(0, index=retornos.index)

        lista = []
        for ticker in retornos.columns:
            if ticker not in MAPA_SETORES: continue
            r = retornos[ticker]
            
            # Cálculos
            ret_total = dm.total_return(r)
            vol = dm.annualize_vol(r, 252)
            sharpe = dm.sharpe_ratio(r, RISK_FREE, 252)
            dd_df = dm.drawdown(r)
            max_dd = dd_df["Drawdown"].min()
            
            downside_r = r[r < 0]
            downside_dev = downside_r.std(ddof=0) * np.sqrt(252)
            ann_ret = dm.annualize_rets(r, 252)
            sortino = (ann_ret - RISK_FREE) / downside_dev if downside_dev != 0 else 0
            
            matrix = np.cov(r, bench_ret)
            beta = matrix[0,1] / matrix[1,1]

            skew = dm.skewness(r)
            kurt = dm.kurtosis(r)
            var_95_normal = dm.var_gaussian(r, level=5, modified=False)
            var_95_hist = dm.var_historic(r, level=5)
            var_95_cf = dm.var_gaussian(r, level=5, modified=True)
            cvar_95 = dm.cvar_historic(r, level=5)
            worst_day = r.min()
            calmar = ann_ret / abs(max_dd) if max_dd != 0 else 0

            lista.append({
                'Ativo': ticker, 'Setor': MAPA_SETORES.get(ticker),
                'Retorno': ret_total, 'Volatilidade': vol,
                'Sharpe': sharpe, 'Sortino': sortino, 'Beta': beta, 'Max DD': max_dd,
                'VaR_Normal': var_95_normal, 'VaR_Hist': var_95_hist, 'VaR_CF': var_95_cf,
                'CVaR': cvar_95, 'Worst Day': worst_day, 'Kurt': kurt, 'Calmar': calmar
            })
        return pd.DataFrame(lista)
    except Exception as e:
        print(f"Erro: {e}")
        return pd.DataFrame()
    finally: conn.close()

# ==============================================================================
# 4. EXCEL BUILDER (VISUAL CORRIGIDO)
# ==============================================================================
def gerar_relatorio_final():
    df = calcular_metricas_sql()
    if df.empty: sys.exit(1)
    print(f"📊 Gerando Relatório Final: {CAMINHO_FINAL}")
    wb = Workbook()
    
    # Estilos
    azul_escuro = "003366"; vermelho_alerta = "C00000"; cinza_claro = "F2F2F2"; laranja = "FFC000"
    header_font = Font(color="FFFFFF", bold=True, size=10)
    fill_azul = PatternFill(start_color=azul_escuro, end_color=azul_escuro, fill_type="solid")
    fill_cinza = PatternFill(start_color=cinza_claro, end_color=cinza_claro, fill_type="solid")
    fill_vermelho = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    center = Alignment(horizontal='center', vertical='center')
    left_indent = Alignment(horizontal='left', vertical='center', indent=1)
    borda = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # === ABA 1: MONITOR ===
    ws1 = wb.active; ws1.title = "Monitor Geral"
    ws1['B2'] = "MONITOR DE PERFORMANCE"; ws1['B2'].font = Font(size=16, bold=True, color=azul_escuro)
    ws1['B3'] = f"Ref: {datetime.now().strftime('%d/%m/%Y')} | RF: {RISK_FREE:.2%}"

    cols1 = ['Ativo', 'Setor', 'Retorno Total', 'Volatilidade', 'Sharpe', 'Sortino', 'Beta', 'Max Drawdown']
    larguras1 = {'B': 12, 'C': 12, 'D': 13, 'E': 13, 'F': 10, 'G': 10, 'H': 9, 'I': 13}
    for col, w in larguras1.items(): ws1.column_dimensions[col].width = w

    for i, c in enumerate(cols1):
        cell = ws1.cell(row=6, column=i+2, value=c)
        cell.fill = fill_azul; cell.font = header_font; cell.alignment = center

    for r_idx, row in enumerate(df.to_dict('records'), 7):
        c_atv = ws1.cell(row=r_idx, column=2, value=row['Ativo'])
        c_atv.font = Font(bold=True); c_atv.alignment = left_indent; c_atv.border = borda
        ws1.cell(row=r_idx, column=3, value=row['Setor']).alignment = center
        
        c_ret = ws1.cell(row=r_idx, column=4, value=row['Retorno'])
        c_ret.number_format = '0.0%'; c_ret.font = Font(color="006100" if row['Retorno']>0 else "9C0006", bold=True)
        
        ws1.cell(row=r_idx, column=5, value=row['Volatilidade']).number_format = '0.0%'
        ws1.cell(row=r_idx, column=6, value=row['Sharpe']).number_format = '0.00'
        ws1.cell(row=r_idx, column=7, value=row['Sortino']).number_format = '0.00'
        ws1.cell(row=r_idx, column=8, value=row['Beta']).number_format = '0.00'
        ws1.cell(row=r_idx, column=9, value=row['Max DD']).number_format = '0.0%'
        
        for c in range(3, 10): ws1.cell(row=r_idx, column=c).border = borda; 
        if r_idx > 6: 
             for c in range(5,10): ws1.cell(row=r_idx, column=c).alignment = center

    ws1.freeze_panes = "B7"
    ultima_linha = 6 + len(df)
    ws1.conditional_formatting.add(f'F7:F{ultima_linha}', DataBarRule(start_type='min', end_type='max', color="638EC6"))
    ws1.conditional_formatting.add(f'E7:E{ultima_linha}', DataBarRule(start_type='min', end_type='max', color="FF99CC"))

    # --- GRÁFICO SCATTER (O CORRETO) ---
    chart1 = ScatterChart()
    chart1.title = "Risco x Retorno (Dispersão)"
    chart1.style = 2; chart1.height = 14; chart1.width = 24
    chart1.x_axis.title = "Volatilidade"; chart1.y_axis.title = "Retorno Total"
    chart1.legend.position = 'r'

    for i in range(len(df)):
        r = 7 + i
        series = Series(
            values=Reference(ws1, min_col=4, min_row=r, max_row=r),
            xvalues=Reference(ws1, min_col=5, min_row=r, max_row=r),
            title_from_data=False, title=ws1.cell(row=r, column=2).value
        )
        series.marker.symbol = "circle"; series.marker.size = 7
        series.graphicalProperties.line.noFill = True # SEM LINHA! (Evita o rabisco)
        chart1.series.append(series)
    ws1.add_chart(chart1, "B18")

    # === ABA 2: BEYOND VAR ===
    ws2 = wb.create_sheet("BeyondVaR Analysis")
    ws2['B2'] = "STRESS TEST (ANÁLISE DE CRISE)"; ws2['B2'].font = Font(size=16, bold=True, color=vermelho_alerta)

    cols2 = ['Ativo', 'VaR Normal', 'VaR Histórico', 'VaR CF (Fat Tail)', 'CVaR (Extreme)', 'Kurtosis', 'Worst Day']
    ws2.column_dimensions['B'].width = 12
    for i in range(2, 9): ws2.column_dimensions[get_column_letter(i+1)].width = 18

    for i, c in enumerate(cols2):
        cell = ws2.cell(row=6, column=i+2, value=c)
        cell.fill = fill_azul; cell.font = header_font; cell.alignment = center

    for r_idx, row in enumerate(df.to_dict('records'), 7):
        c_atv = ws2.cell(row=r_idx, column=2, value=row['Ativo'])
        c_atv.font = Font(bold=True); c_atv.alignment = left_indent; c_atv.border = borda
        ws2.cell(row=r_idx, column=3, value=row['VaR_Normal']).number_format = '0.00%'
        ws2.cell(row=r_idx, column=4, value=row['VaR_Hist']).number_format = '0.00%'
        ws2.cell(row=r_idx, column=5, value=row['VaR_CF']).number_format = '0.00%'
        c_cvar = ws2.cell(row=r_idx, column=6, value=row['CVaR'])
        c_cvar.number_format = '0.00%'; c_cvar.font = Font(bold=True, color=vermelho_alerta)
        c_kurt = ws2.cell(row=r_idx, column=7, value=row['Kurt'])
        c_kurt.number_format = '0.00'
        if row['Kurt'] > 3: c_kurt.font = Font(bold=True, color=laranja)
        c_wd = ws2.cell(row=r_idx, column=8, value=row['Worst Day'])
        c_wd.number_format = '0.00%'; c_wd.font = Font(italic=True)
        for c in range(3, 9): ws2.cell(row=r_idx, column=c).border = borda; ws2.cell(row=r_idx, column=c).alignment = center
    
    ws2.freeze_panes = "B7"

    chart2 = BarChart(); chart2.type = "col"; chart2.style = 10
    chart2.title = "Escada de Risco"; chart2.y_axis.title = "% Perda Estimada"; chart2.height = 14; chart2.width = 30
    data2 = Reference(ws2, min_col=3, min_row=6, max_row=ultima_linha, max_col=6)
    cats2 = Reference(ws2, min_col=2, min_row=7, max_row=ultima_linha)
    chart2.add_data(data2, titles_from_data=True); chart2.set_categories(cats2)
    ws2.add_chart(chart2, "B18")

    # Alerta
    row_alert = 45
    ws2.merge_cells(start_row=row_alert, start_column=2, end_row=row_alert, end_column=4)
    c_alert = ws2.cell(row=row_alert, column=2, value="ALERTA DE RISCO (TOP 3)")
    c_alert.fill = PatternFill(start_color="B30000", end_color="B30000", fill_type="solid"); c_alert.font = header_font; c_alert.alignment = center
    
    for idx, row in enumerate(df.sort_values(by='Kurt', ascending=False).head(3).to_dict('records')):
        r = row_alert + 1 + idx
        ws2.cell(row=r, column=2, value=row['Ativo']).border = borda
        ws2.cell(row=r, column=3, value=f"Kurt: {row['Kurt']:.2f}").border = borda
        c_s = ws2.cell(row=r, column=4, value="CRÍTICO"); c_s.border = borda; c_s.font = Font(color="FF0000", bold=True); c_s.fill = fill_vermelho; c_s.alignment = center

    # --- GLOSSÁRIO RECUPERADO ---
    row_gloss = row_alert + 6
    ws2.cell(row=row_gloss, column=2, value="GLOSSÁRIO TÉCNICO").font = Font(bold=True, color=azul_escuro)
    termos = [
        ("VaR Normal", "Cenário Otimista. Subestima riscos de cauda (Gaussiano)."),
        ("VaR Cornish-Fisher", "Cenário Realista. Ajustado para assimetria e curtose do ativo."),
        ("CVaR (Extreme)", "Cenário de Crise. Média das perdas quando o VaR é rompido."),
        ("Calmar Ratio", "Retorno Anual / Max Drawdown. Eficiência na dor.")
    ]
    for i, (t, d) in enumerate(termos):
        l = row_gloss + 1 + i
        c_t = ws2.cell(row=l, column=2, value=t); c_t.font = Font(bold=True); c_t.border = borda; c_t.fill = fill_cinza
        ws2.merge_cells(start_row=l, start_column=3, end_row=l, end_column=7) # Mais largo
        c_d = ws2.cell(row=l, column=3, value=d); c_d.border = borda
        for c in range(3, 8): ws2.cell(row=l, column=c).border = borda

    wb.save(CAMINHO_FINAL)
    print(f"✅ Relatório Final Gerado: {CAMINHO_FINAL}")

if __name__ == "__main__":
    gerar_relatorio_final()