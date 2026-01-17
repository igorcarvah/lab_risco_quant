import pandas as pd
import numpy as np
from scipy.stats import skew, kurtosis
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.formatting.rule import DataBarRule
import sqlite3
import os
import sys
from datetime import datetime

# ==============================================================================
# 1. CONFIGURAÇÃO
# ==============================================================================
DIRETORIO_ATUAL = os.path.dirname(os.path.abspath(__file__)) 
RAIZ_PROJETO = os.path.dirname(os.path.dirname(DIRETORIO_ATUAL)) 
CAMINHO_DB = os.path.join(RAIZ_PROJETO, "dados", "mercado.db")
PASTA_SAIDA = os.path.join(RAIZ_PROJETO, "reports")
NOME_ARQUIVO = f"Relatorio_Risco_Final_{datetime.now().strftime('%Y%m%d')}.xlsx"
CAMINHO_FINAL = os.path.join(PASTA_SAIDA, NOME_ARQUIVO)

MAPA_SETORES = {
    '^BVSP': 'Benchmark',
    'VALE3.SA': 'Commodity', 'PETR4.SA': 'Commodity',
    'ITUB4.SA': 'Financeiro', 'BPAC11.SA': 'Financeiro',
    'MGLU3.SA': 'Varejo', 'LREN3.SA': 'Varejo',
    'WEGE3.SA': 'Defensiva', 'TAEE11.SA': 'Defensiva'
}
RISK_FREE = 0.1075

if not os.path.exists(PASTA_SAIDA): os.makedirs(PASTA_SAIDA)

# ==============================================================================
# 2. MOTOR DE CÁLCULO
# ==============================================================================
def calcular_metricas_sql():
    if not os.path.exists(CAMINHO_DB): return pd.DataFrame()
    conn = sqlite3.connect(CAMINHO_DB)
    try:
        df_precos = pd.read_sql("SELECT * FROM cotacoes", conn, index_col='Date')
        df_precos.index = pd.to_datetime(df_precos.index)
        retornos = df_precos.pct_change().dropna()
        
        lista = []
        for ticker in retornos.columns:
            if ticker not in MAPA_SETORES: continue
            r = retornos[ticker]
            
            lista.append({
                'Ativo': ticker, 
                'Setor': MAPA_SETORES.get(ticker),
                'Retorno': (1 + r).prod() - 1,
                'Volatilidade': r.std() * np.sqrt(252),
                'Sharpe': (r.mean() * 252 - RISK_FREE) / (r.std() * np.sqrt(252)),
                'VaR 95%': np.percentile(r, 5),
                'Skew': skew(r),
                'Kurt': kurtosis(r), 
                'Max DD': ((1 + r).cumprod() / (1 + r).cumprod().cummax() - 1).min()
            })
        return pd.DataFrame(lista)
    except Exception as e: print(f"Erro: {e}"); return pd.DataFrame()
    finally: conn.close()

# ==============================================================================
# 3. EXCEL BUILDER (PIXEL PERFECT)
# ==============================================================================
df = calcular_metricas_sql()
if df.empty: sys.exit(1)

print(f"📊 Gerando Relatório Compacto em: {CAMINHO_FINAL}")
wb = Workbook()
ws = wb.active
ws.title = "Dashboard Risco"

# --- CONFIGURAÇÃO DE LARGURAS FIXAS (O SEGREDO DO DESIGN) ---
# Definimos na mão para garantir que fique "apertadinho" e elegante
larguras = {
    'B': 10,  # Ativo (Pequeno)
    'C': 12,  # Setor
    'D': 11,  # Retorno
    'E': 11,  # Volatilidade
    'F': 11,  # Sharpe
    'G': 11,  # VaR
    'H': 9,   # Skew (Números pequenos)
    'I': 9,   # Kurt (Números pequenos)
    'J': 11   # Max DD
}
for col, width in larguras.items():
    ws.column_dimensions[col].width = width

# Estilos
azul_escuro = "003366" # Azul Itaú/BTG
vermelho_alerta = "C00000"
cinza_claro = "F2F2F2"
header_font = Font(color="FFFFFF", bold=True, size=10) # Fonte menor no header para caber
fill_azul = PatternFill(start_color=azul_escuro, end_color=azul_escuro, fill_type="solid")
center = Alignment(horizontal='center', vertical='center')
borda = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Cabeçalho Principal
ws['B2'] = "MONITOR DE RISCO QUANTITATIVO"
ws['B2'].font = Font(size=18, bold=True, color=azul_escuro)
ws['B3'] = f"Data: {datetime.now().strftime('%d/%m/%Y')} | RF: {RISK_FREE:.2%}"

# --------------------------------------------------------------------------
# 1. TABELA PRINCIPAL
# --------------------------------------------------------------------------
row_start = 6
cols = ['Ativo', 'Setor', 'Retorno', 'Volat.', 'Sharpe', 'VaR 95%', 'Skew', 'Kurt', 'Max DD'] # Abreviei titulos

# Renderizar Cabeçalhos
for i, c in enumerate(cols):
    cell = ws.cell(row=row_start, column=i+2)
    cell.value = c; cell.fill = fill_azul; cell.font = header_font; cell.alignment = center

# Renderizar Dados
for r_idx, row in enumerate(df.to_dict('records'), row_start + 1):
    # Ativo e Setor
    ws.cell(row=r_idx, column=2, value=row['Ativo']).font = Font(bold=True, size=10)
    ws.cell(row=r_idx, column=3, value=row['Setor']).font = Font(size=10)
    
    # Numéricos (Font Size 10 para ficar delicado)
    c_ret = ws.cell(row=r_idx, column=4, value=row['Retorno'])
    c_ret.number_format = '0.0%'; c_ret.font = Font(color="006100" if row['Retorno'] > 0 else "9C0006", bold=True, size=10)

    ws.cell(row=r_idx, column=5, value=row['Volatilidade']).number_format = '0.0%'
    ws.cell(row=r_idx, column=6, value=row['Sharpe']).number_format = '0.00'
    ws.cell(row=r_idx, column=7, value=row['VaR 95%']).number_format = '0.0%'
    ws.cell(row=r_idx, column=8, value=row['Skew']).number_format = '0.00'
    
    # Destaque Kurtosis
    c_kurt = ws.cell(row=r_idx, column=9, value=row['Kurt'])
    c_kurt.number_format = '0.00'
    if row['Kurt'] > 3: c_kurt.font = Font(bold=True, color=vermelho_alerta, size=10)
    
    # Destaque VaR
    if row['VaR 95%'] < -0.04: ws.cell(row=r_idx, column=7).font = Font(bold=True, color=vermelho_alerta, size=10)

    ws.cell(row=r_idx, column=10, value=row['Max DD']).number_format = '0.0%'
    
    # Aplicar bordas e alinhamento em tudo
    for c in range(2, 11): 
        cell = ws.cell(row=r_idx, column=c)
        cell.border = borda
        cell.alignment = center
        if cell.font.size is None: cell.font = Font(size=10) # Garante fonte 10 se não tiver

ultima_linha = row_start + len(df)

# Barras de Dados (Agora vão aparecer bem porque a coluna é estreita)
ws.conditional_formatting.add(f'E{row_start+1}:E{ultima_linha}', DataBarRule(start_type='min', end_type='max', color="FF6384")) # Vol (Vermelho claro)
ws.conditional_formatting.add(f'F{row_start+1}:F{ultima_linha}', DataBarRule(start_type='min', end_type='max', color="36A2EB")) # Sharpe (Azul claro)

# --------------------------------------------------------------------------
# 2. GRÁFICO (CENTRALIZADO EMBAIXO)
# --------------------------------------------------------------------------
linha_grafico = ultima_linha + 3
chart = ScatterChart()
chart.title = "Risco x Retorno"
chart.style = 2; chart.height = 10; chart.width = 16 # Menor para acompanhar a largura da tabela

xval = Reference(ws, min_col=5, min_row=row_start+1, max_row=ultima_linha) # Vol
yval = Reference(ws, min_col=4, min_row=row_start+1, max_row=ultima_linha) # Retorno
series = Series(yval, xval, title_from_data=False)
series.marker.symbol = "circle"; series.marker.graphicalProperties.solidFill = azul_escuro
chart.series.append(series)
ws.add_chart(chart, f"B{linha_grafico}") # Coluna B

# --------------------------------------------------------------------------
# 3. ALERTA (EMBAIXO DO GRÁFICO)
# --------------------------------------------------------------------------
linha_alerta = linha_grafico + 21
col_alerta = 2

ws.merge_cells(start_row=linha_alerta, start_column=col_alerta, end_row=linha_alerta, end_column=col_alerta+2) # B, C, D
ws.cell(row=linha_alerta, column=col_alerta, value="⚠️ ALERTA DE RISCO (TOP 3)").fill = PatternFill(start_color="B30000", end_color="B30000", fill_type="solid")
ws.cell(row=linha_alerta, column=col_alerta).font = Font(color="FFFFFF", bold=True, size=10); ws.cell(row=linha_alerta, column=col_alerta).alignment = center

top3 = df.sort_values(by='Kurt', ascending=False).head(3)
for idx, row in enumerate(top3.to_dict('records')):
    r = linha_alerta + 1 + idx
    # Ticker
    ws.cell(row=r, column=col_alerta, value=row['Ativo']).border = borda; ws.cell(row=r, column=col_alerta).alignment = center
    # Valor
    c_v = ws.cell(row=r, column=col_alerta+1, value=f"{row['Kurt']:.2f}")
    c_v.border = borda; c_v.alignment = center
    # Status
    c_s = ws.cell(row=r, column=col_alerta+2, value="CRÍTICO")
    c_s.font = Font(color="FF0000", bold=True, size=9); c_s.border = borda; c_s.alignment = center

# --------------------------------------------------------------------------
# 4. GLOSSÁRIO 
# --------------------------------------------------------------------------
linha_gloss = linha_alerta + 5
ws.cell(row=linha_gloss, column=2, value="GLOSSÁRIO").font = Font(bold=True, color=azul_escuro, size=10)

# Lista Definitiva de Definições
termos = [
    ("Retorno", "Rentabilidade acumulada no período analisado."),
    ("Volatilidade", "Incerteza do preço (Desvio Padrão Anualizado). Mede a intensidade do 'sobe e desce'."),
    ("Sharpe Ratio", "Eficiência. Quanto o ativo pagou de retorno para cada unidade de risco corrido (acima do CDI)."),
    ("VaR 95%", "Value at Risk. Com 95% de confiança, a perda diária não excederá este valor."),
    ("Skewness", "Assimetria. Negativo = Viés de queda (Cai rápido, sobe devagar)."),
    ("Kurtosis", "Detector de Caudas Gordas. > 3 indica alta probabilidade de eventos extremos (Crises)."),
    ("Max DD", "Max Drawdown. A pior queda percentual registrada do topo histórico até o fundo.")
]

for i, (t, d) in enumerate(termos):
    l = linha_gloss + 1 + i
    
    # Coluna B: Nome do Termo (Fundo cinza)
    c_t = ws.cell(row=l, column=2, value=t)
    c_t.font = Font(bold=True, size=9); c_t.border = borda
    c_t.fill = PatternFill(start_color=cinza_claro, end_color=cinza_claro, fill_type="solid")
    
    # Coluna C até J: Descrição (Mesclado)
    ws.merge_cells(start_row=l, start_column=3, end_row=l, end_column=10)
    c_d = ws.cell(row=l, column=3, value=d)
    c_d.font = Font(size=9); c_d.alignment = Alignment(horizontal='left')
    
    # Borda na área mesclada
    for c in range(3, 11): ws.cell(row=l, column=c).border = borda

wb.save(CAMINHO_FINAL)
print(f"✅ Relatório 'Pixel Perfect' gerado: {CAMINHO_FINAL}")

# ==============================================================================
# 4. EXECUÇÃO SEGURA
# ==============================================================================
if __name__ == "__main__":
    main()